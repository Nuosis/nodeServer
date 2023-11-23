import os
import uuid
import openpyxl
import json
import requests
import io
import logging
import importlib.util
import argparse
from datetime import datetime

# Argparse setup
parser = argparse.ArgumentParser(description='Convert XLSX to JSON.')
parser.add_argument('url', type=str, help='URL of the XLSX file')
parser.add_argument('formName', type=str, nargs='?', default=None, help='Name of the form (optional)')
args = parser.parse_args()

# Configure logging
#DEBUG, INFO, WARNING, ERROR, and CRITICAL.
log_file_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), 'error_log.log')
logging.basicConfig(filename=log_file_path, level=logging.DEBUG, 
                    format='%(asctime)s:%(levelname)s:%(message)s')

# Get the directory of the script
script_directory = os.path.dirname(os.path.abspath(__file__))

# Define the base directory by appending 'images' to the script directory
base_directory = os.path.join(script_directory, 'images')

def save_images_from_sheet(sheet, image_directory):
    # Combine the base directory with the specific image directory
    full_image_directory = os.path.join(base_directory, image_directory)
    #print(f"Full Image Directory: {full_image_directory}")
    #logging.error(f"Full Image Directory: {full_image_directory}")

    if not os.path.exists(full_image_directory):
        os.makedirs(full_image_directory)

    image_paths = []
    for index, drawing in enumerate(sheet._images):
        # [rest of your code for handling the image data]

        # Adjust the image path to use the full directory path
        image_filename = f"{sheet.title}_image_{index}.png"
        image_path = os.path.join(full_image_directory, image_filename)

        with open(image_path, "wb") as img_file:
            img_file.write(image_data)

        # Adjust the image_paths to reflect the new location
        image_paths.append(os.path.join('/images', image_directory, image_filename))

    return image_paths



def xlsx_to_json_from_url(url, formName=None):
    try:
        # Attempt to download the file content
        response = requests.get(url)
        response.raise_for_status()
    except requests.exceptions.HTTPError as http_err:
        logging.error(f"HTTP Error occurred: {http_err}")
        print(json.dumps({"error": f"HTTP Error occurred: {http_err}"}))
        return
    except requests.exceptions.ConnectionError as conn_err:
        logging.error(f"Connection Error occurred: {conn_err}")
        print(json.dumps({"error": f"Connection Error occurred: {conn_err}"}))
        return
    except requests.exceptions.Timeout as timeout_err:
        logging.error(f"Timeout Error occurred: {timeout_err}")
        print(json.dumps({"error": f"Timeout Error occurred: {timeout_err}"}))
        return
    except requests.exceptions.RequestException as req_err:
        logging.error(f"Error occurred while making the request: {req_err}")
        print(json.dumps({"error": f"Error occurred while making the request: {req_err}"}))
        return

    try:
        # Use BytesIO to handle the downloaded content
        file_content = io.BytesIO(response.content)
        # Load workbook from the in-memory bytes
        workbook = openpyxl.load_workbook(file_content, data_only=True)
    except openpyxl.utils.exceptions.InvalidFileException:
        logging.error("The downloaded file is not a valid Excel file.")
        print(json.dumps({"error": "The downloaded file is not a valid Excel file."}))
        return
    except Exception as e:
        logging.error(f"An error occurred while processing the Excel file: {e}")
        print(json.dumps({"error": f"An error occurred while processing the Excel file: {e}"}))
        return
        
    if formName:
        # Get the directory where the script is located
        script_dir = os.path.dirname(os.path.realpath(__file__))

        # Construct the path to the form module
        module_name = f"{formName}.py"
        module_path = os.path.join(script_dir, module_name)

        if os.path.isfile(module_path):
            spec = importlib.util.spec_from_file_location(formName, module_path)
            form_module = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(form_module)

            # Execute the module's function and store the result
            extracted_data = form_module.extract_data(workbook)

           # Log the result
            logging.info(f"Result from '{formName}' form: {extracted_data}")
            print(json.dumps({"data": extracted_data}))
            return
        else:
            logging.error("Form not set up")
            print(json.dumps({"error": "Form not set up"}))
            return
    
    # Existing code for processing without form-specific file
    all_sheets_data = {}
    image_uuid = uuid.uuid4().hex[:10]
    image_directory = os.path.join('images', image_uuid)

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        data = []

        keys = [cell.value for cell in sheet[1]]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = {}
            for key, value in zip(keys, row):
                if isinstance(value, datetime):
                    # Convert datetime to string
                    row_data[key] = value.isoformat()
                else:
                    row_data[key] = value
            data.append(row_data)

        image_paths = save_images_from_sheet(sheet, image_directory)
        all_sheets_data[sheet_name] = {'data': data, 'images': image_paths}
        
    extracted_data = json.dumps(all_sheets_data, indent=4)

    # Log a confirmation that the default extraction was completed
    logging.info("Default data extraction completed successfully.")

    print(extracted_data)  # Print the final extracted data in JSON format

if __name__ == "__main__":
    result = xlsx_to_json_from_url(args.url, args.formName)
    #print(result)
